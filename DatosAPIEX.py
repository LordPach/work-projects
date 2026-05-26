import requests
import time
from datetime import datetime
from openpyxl import load_workbook

# ================= CONFIGURACIÓN =================
BASE_URL = "Your api link here"
TOKEN = "your api token here."
HEADERS = {"Authorization": f"Bearer {TOKEN}", "Content-Type": "application/json"}

EXCEL_FILE = "datos.xlsx"
LOG_FILE = f"log_tickets_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"


# ================= LOG =================
def log(message, level="INFO"):
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    formatted = f"[{level}] {timestamp} - {message}"
    print(formatted)
    with open(LOG_FILE, "a", encoding="utf-8") as f:
        f.write(formatted + "\n")


# ================= EXCEL =================
def read_excel_data():
    try:
        log("📂 Leyendo Excel...")
        wb = load_workbook(EXCEL_FILE)
        sheet = wb.active

        data = []

        for row in sheet.iter_rows(min_row=2, max_col=2, values_only=True):
            field_id, value = row

            if not field_id:
                continue

            data.append({
                "CustomField_id": str(field_id).strip(),
                "Value": str(value) if value is not None else ""
            })

        log(f"📊 Campos cargados: {len(data)}")
        return data

    except Exception as e:
        log(f"⛔ Error leyendo Excel: {e}", "ERROR")
        return None


# ================= API =================
def get_all_incidents():
    try:
        log("🔄 Obteniendo tickets...")
        response = requests.get(f"{BASE_URL}/Incidents", headers=HEADERS, timeout=15)
        response.raise_for_status()
        data = response.json()

        if isinstance(data, dict) and "Items" in data:
            return data["Items"]
        elif isinstance(data, list):
            return data

        return []

    except Exception as e:
        log(f"⛔ Error obteniendo tickets: {e}", "ERROR")
        return None


def get_ticket_custom_fields(ticket_id):
    try:
        response = requests.get(
            f"{BASE_URL}/Incidents/{ticket_id}/customFields",
            headers=HEADERS,
            timeout=10
        )
        response.raise_for_status()
        return response.json()
    except Exception as e:
        log(f"⛔ Error obteniendo campos del ticket {ticket_id}: {e}", "ERROR")
        return None


def update_ticket_custom_fields(ticket_id, fields):
    try:
        response = requests.put(
            f"{BASE_URL}/Incidents/{ticket_id}/customFields",
            headers=HEADERS,
            json=fields,
            timeout=10
        )

        log(f"📥 Respuesta API ({ticket_id}): {response.status_code} → {response.text}")

        if response.status_code in (200, 204):
            return True, response.text

        return False, response.text

    except Exception as e:
        return False, str(e)


# ================= MAIN =================
def main():
    log("===== INICIO =====")

    excel_data = read_excel_data()
    if not excel_data:
        log("⛔ Excel inválido", "ERROR")
        return

    tickets = get_all_incidents()
    if tickets is None:
        log("⛔ No se pudieron obtener tickets", "ERROR")
        return

    log(f"📊 Tickets obtenidos: {len(tickets)}")

    evaluated = 0
    updated = 0
    skipped = 0
    failed = 0

    for ticket in tickets:

        evaluated += 1

        ticket_id = ticket.get("Id")
        archived = ticket.get("Archived", False)

        if not ticket_id:
            continue

        if archived:
            log(f"📦 Ticket archivado ignorado: {ticket_id}")
            continue

        ticket_fields = get_ticket_custom_fields(ticket_id)

        if ticket_fields is None:
            failed += 1
            continue

        valid_updates = []

        for excel_field in excel_data:

            match = next(
                (f for f in ticket_fields
                 if f.get("CustomField_id") == excel_field["CustomField_id"]),
                None
            )

            if not match:
                continue

            current_value = match.get("Value")

            # 🔥 SOLO si está vacío
            if current_value:
                log(f"⏭ Ticket {ticket_id} campo {excel_field['CustomField_id']} ya tiene valor → {current_value}")
                continue

            log(f"📝 Ticket {ticket_id} → {excel_field['CustomField_id']} = {excel_field['Value']}")

            valid_updates.append(excel_field)

        if not valid_updates:
            skipped += 1
            continue

        success, msg = update_ticket_custom_fields(ticket_id, valid_updates)

        if success:
            updated += 1
            log(f"✅ Ticket actualizado: {ticket_id}")
        else:
            failed += 1
            log(f"❌ Error en {ticket_id}: {msg}", "ERROR")

        # evitar rate limit
        time.sleep(0.2)

        if evaluated % 50 == 0:
            log(f"📊 Evaluados: {evaluated} | Actualizados: {updated} | Fallos: {failed}")

    # ================= RESUMEN =================
    log("\n===== RESULTADO FINAL =====")
    log(f"Evaluados: {evaluated}")
    log(f"Actualizados: {updated}")
    log(f"Omitidos: {skipped}")
    log(f"Fallos: {failed}")

    log("===== FIN =====")


if __name__ == "__main__":
    main()
